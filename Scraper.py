#Author:    Ryan H Lee
#Email:     rhl43@rutgers.edu
#
#
#README
#Identifies antibody-based drugs within an Excel spreadsheet, highlights the rows,
#then searches the web for chain sequences and adds them to the spreadsheet
#
# *** Important ***
# MUST install required libraries: openpyxl, pandas, requests, beautifulsoup
# !pip install openpyxl
# !pip install pandas
# !pip install requests
# !pip install beautifulsoup4
#


import os
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, NamedStyle
from openpyxl.utils import get_column_letter
from bs4 import BeautifulSoup
import pandas as pd
import requests

#Import spreadsheet and select worksheet
wb = load_workbook('drugdata.xlsm')
ws = wb.active

#Create an Excel NamedStyle, which will be applied to the cells in the rows and add style to workbook
highlight = NamedStyle(name = 'fill_red', fill = PatternFill(fill_type = 'solid', fgColor = "FF0000"), font = Font(name = 'Calibri', size = 12, color = '000000'))
wb.add_named_style(highlight)

#Iterate through cells in column C, to search for names ending in 'mab'
#Some columns coincidentally have 'mab' appear accidentally, searching through only column C prevents unintentional highlighting

drugs = []
highlight_rows = []

for row_index, row_data in enumerate(ws[f"{'C'}{'0'}:{'C'}{'14595'}"]):
    search_term_1 = 'mab'
    search_term_2 = 'Mab'
    search_term_3 = 'mAb'
    search_term_4 = 'cept'


    #If a name is found with 'mab' or 'cept' suffix, add that row to highlight_rows list
    for cell in row_data:
      if cell.value:
        if search_term_1 in str(cell.value):
          highlight_rows.append(str(cell.row))
          drugs.append(str(ws['A' + str(cell.row)].value))
        elif search_term_2 in str(cell.value):
          highlight_rows.append(str(cell.row))
          drugs.append(str(ws['A' + str(cell.row)].value))
        elif search_term_3 in str(cell.value):
          highlight_rows.append(str(cell.row))
          drugs.append(str(ws['A' + str(cell.row)].value))
        elif search_term_4 in str(cell.value):
          highlight_rows.append(str(cell.row))
          drugs.append(str(ws['A' + str(cell.row)].value))


#For each cell in the row in the highlight_rows list, highlight the cell
for row_str in highlight_rows:
  for cell in ws[f"{row_str}:{row_str}"]:
    cell.style = highlight.name


chains = []
urls = []

for drug in drugs:
  url = "https://go.drugbank.com/drugs/" + drug
  urls.append(str(url))
  page = requests.get(url)
  soup = BeautifulSoup(page.content, "html.parser")
  chain = soup.find_all('pre', attrs={'class':'sequence bg-light'})
  sequence = []
  for sequences in chain:
    sequence.append(sequences.text)
  chains.append(sequence)


i = 0

for row_index, row_data in enumerate(ws[f"{'A'}{'0'}:{'A'}{'14595'}"]):
  for cell in row_data:
    if cell.style == highlight.name:
      ws['H' + str(cell.row)] = str(chains[i])
      ws['I' + str(cell.row)] = str(urls[i])
      i += 1


wb.save('highlightedAndSequenced.xlsx')


df = pd.DataFrame({'Drug':drugs,'Light Chain':chains})
df.to_csv('drugs.csv', index = False, encoding='utf-8')
